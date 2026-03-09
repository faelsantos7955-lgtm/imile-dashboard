"""
excel_export.py — Geração do arquivo Excel formatado
Mesmo formato do app_expedicao.py original
"""
import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import Outline

# ── Estilos ───────────────────────────────────────────────────
_HF   = PatternFill("solid", fgColor="1F3864")
_AF   = PatternFill("solid", fgColor="D9E1F2")
_RF   = PatternFill("solid", fgColor="FF0000")
_GF   = PatternFill("solid", fgColor="70AD47")
_PF   = PatternFill("solid", fgColor="5B9BD5")
_HFNT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_BFNT = Font(name="Arial", size=10)
_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LFT  = Alignment(horizontal="left",   vertical="center")
_TH   = Side(style="thin", color="BFBFBF")
_BRD  = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)

def _titulo_aba(ws, titulo, n):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n, 1))
    c = ws.cell(row=1, column=1, value=titulo)
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = _HF
    c.alignment = _CTR
    ws.row_dimensions[1].height = 30

def _escrever_tabela(ws, df, li=3):
    for ci, cn in enumerate(df.columns, 1):
        c = ws.cell(row=li-1, column=ci, value=cn)
        c.font, c.fill, c.alignment, c.border = _HFNT, _HF, _CTR, _BRD
    for ri, row in enumerate(df.itertuples(index=False), li):
        alt = ri % 2 == 0
        for ci, val in enumerate(row, 1):
            try:
                if pd.isnull(val): val = None
            except: pass
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _BFNT; c.alignment = _LFT; c.border = _BRD
            if alt: c.fill = _AF
            cn = df.columns[ci-1]
            if cn in ("Taxa de Expedicao", "Taxa Exp.", "taxa_exp"):
                c.number_format = "0.0%"
                if isinstance(val, (int, float)) and val < 0.5:
                    c.fill = _RF
                    c.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
            elif cn in ("recebido no DS", "em rota de entrega", "Total Geral",
                        "recebido", "expedido", "entregas"):
                c.number_format = "#,##0"; c.alignment = _CTR

def _auto_width(ws, extra=4):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        mx = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[cl].width = mx + extra

def _escrever_tabela_agrupada(ws, df_pivot, df_cidades, li=3):
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)
    COLS = ["Scan Station", "recebido no DS", "em rota de entrega", "Total Geral",
            "Taxa de Expedicao", "Entregas", "Taxa de Entrega"]
    DS_F  = PatternFill("solid", fgColor="BDD7EE")
    CF1   = PatternFill("solid", fgColor="F2F2F2")
    CF2   = PatternFill("solid", fgColor="FFFFFF")
    DF    = Font(name="Arial", size=10, bold=True,   color="1F3864")
    CITF  = Font(name="Arial", size=9,  italic=True, color="2E75B6")

    # índice cidade por DS
    city_index = {}
    if df_cidades is not None and len(df_cidades) > 0:
        for ds, grp in df_cidades.groupby("Scan Station", observed=True):
            city_index[ds] = grp.sort_values("Total Geral", ascending=False)

    # cabeçalho
    for ci, cn in enumerate(COLS, 1):
        c = ws.cell(row=li-1, column=ci, value=cn)
        c.font, c.fill, c.alignment, c.border = _HFNT, _HF, _CTR, _BRD

    cur = li
    for _, dr in df_pivot.iterrows():
        ds       = dr["Scan Station"]
        ent_ds   = int(dr.get("Entregas", 0) or 0)
        taxa_e   = float(dr.get("Taxa de Entrega", 0) or 0)
        meta     = float(dr.get("Meta", 0.5) or 0.5)
        taxa_exp = float(dr.get("Taxa de Expedicao", 0) or 0)
        vals = [ds,
                dr.get("recebido no DS", 0),
                dr.get("em rota de entrega", 0),
                dr.get("Total Geral", 0),
                taxa_exp, ent_ds, taxa_e]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=cur, column=ci, value=val)
            c.font, c.fill, c.border = DF, DS_F, _BRD
            c.alignment = _LFT if ci == 1 else _CTR
            cn = COLS[ci-1]
            if cn == "Taxa de Expedicao":
                c.number_format = "0.0%"
                c.fill = _RF if taxa_exp < meta else _GF
                c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            elif cn == "Taxa de Entrega":
                c.number_format = "0.0%"; c.fill = _PF
                c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            elif cn in ("recebido no DS", "em rota de entrega", "Total Geral", "Entregas"):
                c.number_format = "#,##0"
        ws.row_dimensions[cur].outline_level = 0
        cur += 1

        # linhas de cidade
        for i, (_, cr) in enumerate(city_index.get(ds, pd.DataFrame()).iterrows()):
            city     = cr.get("Destination City", "")
            fill     = CF1 if i % 2 == 0 else CF2
            ent_c    = int(cr.get("Entregas", 0) or 0)
            taxa_e_c = float(cr.get("Taxa de Entrega", 0) or 0)
            taxa_x_c = float(cr.get("Taxa de Expedicao", 0) or 0)
            vals2 = [f"   {city}",
                     cr.get("recebido no DS", 0),
                     cr.get("em rota de entrega", 0),
                     cr.get("Total Geral", 0),
                     taxa_x_c, ent_c, taxa_e_c]
            for ci, val in enumerate(vals2, 1):
                c = ws.cell(row=cur, column=ci, value=val)
                c.font, c.fill, c.border = CITF, fill, _BRD
                c.alignment = _LFT if ci == 1 else _CTR
                cn = COLS[ci-1]
                if cn == "Taxa de Expedicao":
                    c.number_format = "0.0%"
                elif cn == "Taxa de Entrega":
                    c.number_format = "0.0%"
                    if taxa_e_c > 0:
                        c.fill = _PF
                        c.font = Font(name="Arial", size=9, italic=True, color="FFFFFF")
                elif cn in ("recebido no DS", "em rota de entrega", "Total Geral", "Entregas"):
                    c.number_format = "#,##0"
            ws.row_dimensions[cur].outline_level = 1
            ws.row_dimensions[cur].hidden = True
            cur += 1

# ══════════════════════════════════════════════════════════════
#  FUNÇÃO PRINCIPAL
# ══════════════════════════════════════════════════════════════
def exportar_excel_bytes(base: pd.DataFrame,
                          geral: pd.DataFrame,
                          capital: pd.DataFrame,
                          metro: pd.DataFrame,
                          country: pd.DataFrame,
                          pivot_cidades: pd.DataFrame = None,
                          data_str: str = "") -> bytes:
    """Gera o Excel completo e retorna bytes para download."""
    wb = Workbook()

    # ── Aba Base Tratada ──────────────────────────────────────
    ws = wb.active
    ws.title = "Base_Tratada"
    titulo = f"Base Tratada{' — ' + data_str if data_str else ''}"
    _titulo_aba(ws, titulo, len(base.columns))
    _escrever_tabela(ws, base, li=3)
    _auto_width(ws)
    ws.freeze_panes = ws.cell(row=4, column=1)

    # ── Abas por região ───────────────────────────────────────
    abas = [
        ("Consolidado_Geral", geral,    f"Consolidado Geral{' — ' + data_str if data_str else ''}"),
        ("Capital",           capital,  "Capital"),
        ("Metropolitan",      metro,    "Metropolitan"),
        ("Countryside",       country,  "Countryside"),
    ]
    for sheet_name, df_p, titulo_aba in abas:
        ws = wb.create_sheet(title=sheet_name)
        _titulo_aba(ws, titulo_aba, 7)
        if pivot_cidades is not None and len(df_p) > 0:
            pc = pivot_cidades[pivot_cidades["Scan Station"].isin(df_p["Scan Station"])]
            _escrever_tabela_agrupada(ws, df_p, pc, li=3)
        else:
            _escrever_tabela(ws, df_p, li=3)
        for col, w in zip([1, 2, 3, 4, 5, 6, 7], [26, 20, 22, 14, 18, 14, 16]):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = ws.cell(row=4, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

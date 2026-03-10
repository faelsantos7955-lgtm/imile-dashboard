"""
modulos/triagem.py — Triagem Analyzer (versão Streamlit)
Análise de Erros de Triagem DC × DS
"""
import io
import os
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ── Constantes de colunas ─────────────────────────────────────
COL_WB   = "Waybill No."
COL_LC   = "Loading station"
COL_DST  = "Destination Statio"
COL_DEL  = "Delivery Station"
COL_TIPO = "Tipo"


# ══════════════════════════════════════════════════════════════
#  LÓGICA DE ANÁLISE (igual ao original)
# ══════════════════════════════════════════════════════════════
def _limpar(df):
    df.columns = df.columns.str.strip()
    for c in df.select_dtypes("object").columns:
        df[c] = df[c].str.strip()
    return df


def _ler_bases(file_obj, log_cb):
    nome = getattr(file_obj, "name", "bases.xlsx")
    log_cb(f"📋 Arquivo Bases: {os.path.basename(nome)}")

    xl     = pd.ExcelFile(file_obj)
    sheets = xl.sheet_names
    aba    = next((s for s in sheets
                   if "base erro" in s.lower() or "erro exp" in s.lower()), sheets[0])
    log_cb(f"   ℹ Aba: '{aba}'")

    df = _limpar(pd.read_excel(file_obj, sheet_name=aba, dtype=str))

    def get_col(names):
        for n in names:
            for c in df.columns:
                if str(c).strip().upper() == n.upper():
                    return c
        return None

    col_base     = get_col(["BASE"])
    col_base_pai = get_col(["BASE_PAI"])
    col_sup      = get_col(["SUPERVISOR"])

    for col, nome_col in [(col_base, "BASE"), (col_base_pai, "BASE_PAI")]:
        if not col:
            raise ValueError(f"Coluna '{nome_col}' não encontrada. Colunas: {list(df.columns)}")

    df = df.dropna(subset=[col_base])
    df = df[df[col_base].str.strip() != ""]

    df_ref = df[[col_base, col_base_pai] + ([col_sup] if col_sup else [])].copy()
    df_ref = df_ref.rename(columns={col_base: "__BASE__", col_base_pai: "__BASE_PAI__"})
    if col_sup:
        df_ref = df_ref.rename(columns={col_sup: "__SUPERVISOR__"})

    log_cb(f"   ✔ {len(df_ref):,} pares BASE→BASE_PAI mapeados")
    return df_ref, bool(col_sup)


def _ler_loading_scan(file_obj, log_cb):
    nome = getattr(file_obj, "name", "scan.xlsx")
    log_cb(f"📄 {os.path.basename(nome)}")

    xl     = pd.ExcelFile(file_obj)
    sheets = xl.sheet_names
    aba    = None

    for s in sheets:
        try:
            tmp  = _limpar(pd.read_excel(file_obj, sheet_name=s, dtype=str, nrows=3))
            cols = [c.lower() for c in tmp.columns]
            if any("waybill" in c for c in cols) and any("destination" in c for c in cols):
                aba = s
                break
        except Exception:
            continue

    if not aba:
        aba = sheets[0]

    # Lê só as colunas necessárias para economizar memória
    try:
        df = _limpar(pd.read_excel(file_obj, sheet_name=aba, dtype=str,
                                    usecols=lambda c: any(k in str(c).lower()
                                    for k in ["waybill","loading","destination","delivery","tipo","type"])))
    except Exception:
        df = _limpar(pd.read_excel(file_obj, sheet_name=aba, dtype=str))

    for col in [COL_WB, COL_LC, COL_DST, COL_DEL]:
        if col not in df.columns:
            raise ValueError(f"Coluna '{col}' não encontrada em {os.path.basename(nome)}.\nColunas: {list(df.columns)}")

    df["__arquivo__"] = os.path.basename(nome)
    log_cb(f"   ✔ {len(df):,} registros  |  aba: '{aba}'")
    return df


def run_analysis(scan_files, bases_file, log_cb, progress_cb):
    """
    Executa análise completa.
    Retorna (ok: bool, resultado: dict, erro: str)
    """
    try:
        n = len(scan_files)
        log_cb(f"📂 {n} LoadingScan(s) + arquivo Bases")
        progress_cb(3)

        # 1. Bases
        log_cb("─" * 40)
        df_ref, tem_supervisor = _ler_bases(bases_file, log_cb)
        progress_cb(10)

        # 2. LoadingScans
        log_cb("─" * 40)
        frames = []
        for i, fp in enumerate(scan_files):
            progress_cb(int(10 + (i / n) * 25))
            frames.append(_ler_loading_scan(fp, log_cb))

        log_cb("─" * 40)
        log_cb(f"🔀 Consolidando {n} arquivo(s)…")
        df = pd.concat(frames, ignore_index=True)
        log_cb(f"✔ {len(df):,} registros no total")
        progress_cb(38)

        # Converte colunas com valores repetidos para category (economiza ~60% de RAM)
        for col in [COL_LC, COL_DST, COL_DEL, "__arquivo__"]:
            if col in df.columns:
                df[col] = df[col].astype("category")

        # 3. OUT BOUND vetorizado
        log_cb("🔍 Calculando OUT BOUND…")
        df = df.merge(
            df_ref[["__BASE__","__BASE_PAI__"]].rename(
                columns={"__BASE__": COL_DST, "__BASE_PAI__": "__BASE_PAI_DST__"}),
            on=COL_DST, how="left"
        )
        conditions = [
            df["__BASE_PAI_DST__"].isna(),
            df["__BASE_PAI_DST__"].str.strip() == df[COL_DEL].str.strip(),
        ]
        choices = ["FORA ABRANGÊNCIA", "OK"]
        df["OUT BOUND"] = np.select(conditions, choices, default="ERRO EXPEDIÇÃO")
        progress_cb(52)

        # 4. Supervisor
        if tem_supervisor:
            log_cb("🔗 Mapeando Supervisores…")
            df = df.merge(
                df_ref[["__BASE__","__SUPERVISOR__"]].rename(
                    columns={"__BASE__": COL_DST, "__SUPERVISOR__": "SUPERVISOR"}),
                on=COL_DST, how="left"
            )
        progress_cb(58)

        total    = len(df)
        qtd_ok   = int((df["OUT BOUND"] == "OK").sum())
        qtd_erro = int((df["OUT BOUND"] == "ERRO EXPEDIÇÃO").sum())
        qtd_fora = int((df["OUT BOUND"] == "FORA ABRANGÊNCIA").sum())
        taxa     = round(qtd_ok / total * 100, 1) if total else 0
        log_cb(f"✔ OK: {qtd_ok:,}  |  Erro: {qtd_erro:,}  |  Fora: {qtd_fora:,}  |  Taxa: {taxa}%")

        # 5. Agregações vetorizadas (muito mais rápido que apply)
        log_cb("📊 Agregando relatórios…")

        # Cria colunas binárias vetorizadas
        df["_ok"] = (df["OUT BOUND"] == "OK").astype("int8")
        df["_er"] = (df["OUT BOUND"] == "ERRO EXPEDIÇÃO").astype("int8")
        df["_fa"] = (df["OUT BOUND"] == "FORA ABRANGÊNCIA").astype("int8")

        def agg_rapida(group_col, rename=None):
            """Agrega métricas por coluna usando soma vetorizada."""
            g = df.groupby(group_col, dropna=False).agg(
                Total_Expedido=(COL_WB, "count"),
                Triagem_OK=("_ok", "sum"),
                Triagem_NOK=("_er", "sum"),
                Fora_Abrangencia=("_fa", "sum"),
            ).reset_index()
            g.columns = [group_col] + ["Total Expedido","Triagem OK","Triagem NOK","Fora Abrangência"]
            g["Taxa (%)"] = (g["Triagem OK"] / g["Total Expedido"].replace(0, np.nan) * 100).round(2).fillna(0)
            if rename:
                g = g.rename(columns={group_col: rename})
            return g

        r_dc  = agg_rapida(COL_LC)
        r_arq = agg_rapida("__arquivo__", rename="Arquivo")
        r_sup = pd.DataFrame()
        if tem_supervisor and "SUPERVISOR" in df.columns:
            r_sup = agg_rapida("SUPERVISOR")
        r_tipo = pd.DataFrame()
        if COL_TIPO in df.columns:
            r_tipo = agg_rapida(COL_TIPO)

        # Remove colunas auxiliares
        df.drop(columns=["_ok","_er","_fa"], inplace=True, errors="ignore")

        df_erro  = df[df["OUT BOUND"] == "ERRO EXPEDIÇÃO"]
        top5_ds  = (df_erro.groupby(COL_DST, dropna=False)[COL_WB]
                    .count().nlargest(5).reset_index()
                    .rename(columns={COL_WB: "Total Erros", COL_DST: "DS"}))
        progress_cb(72)

        # 6. Gera Excel
        log_cb("💾 Gerando Excel…")
        excel_bytes = _gerar_excel(
            df, r_dc, r_arq, r_sup, r_tipo, top5_ds,
            qtd_ok, qtd_erro, qtd_fora, taxa, log_cb
        )
        progress_cb(100)
        log_cb("✅ Concluído!")

        return True, {
            "total": total, "ok": qtd_ok, "erro": qtd_erro,
            "fora": qtd_fora, "taxa": taxa,
            "r_dc": r_dc, "r_arq": r_arq, "r_sup": r_sup,
            "top5": top5_ds, "excel_bytes": excel_bytes,
            "df_full": df,
        }, ""

    except Exception:
        import traceback
        return False, {}, traceback.format_exc()


# ══════════════════════════════════════════════════════════════
#  GERAÇÃO DO EXCEL
# ══════════════════════════════════════════════════════════════
def _gerar_excel(df, r_dc, r_arq, r_sup, r_tipo, top5_ds,
                  qtd_ok, qtd_erro, qtd_fora, taxa, log_cb):
    AZUL_ESC = "0D1F6E"; AZUL_MED = "1F4FEB"
    VERM     = "C00000"; VERD     = "1E6B3C"
    CINZA    = "F2F2F2"; BRANCO   = "FFFFFF"

    hdr_fill = PatternFill("solid", fgColor=AZUL_ESC)
    hdr_font = Font(bold=True, color=BRANCO, name="Calibri", size=10)
    ok_fill  = PatternFill("solid", fgColor="D6F0DC")
    nok_fill = PatternFill("solid", fgColor="FFE0E0")
    thin     = Side(style="thin", color="CCCCCC")
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def auto_w(ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 32)

    def fmt_hdr(ws):
        for cell in ws[1]:
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ctr
        ws.freeze_panes = "A2"

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    # ── Helper: escreve DataFrame em worksheet openpyxl ──────
    def df_to_ws(ws, df_in, start_row=1):
        """Escreve DataFrame em worksheet usando append() — muito mais rápido que iterrows."""
        import numpy as np
        # Converte category para str antes de escrever
        df_in = df_in.copy()
        for col in df_in.select_dtypes(include="category").columns:
            df_in[col] = df_in[col].astype(str)
        # Move para a linha correta se start_row > 1
        for _ in range(start_row - 1):
            ws.append([])
        # Header
        ws.append(list(df_in.columns))
        # Dados: converte numpy types para Python nativo
        for row in df_in.itertuples(index=False, name=None):
            ws.append([
                None if (isinstance(v, float) and np.isnan(v))
                else int(v) if isinstance(v, (np.integer,))
                else float(v) if isinstance(v, (np.floating,))
                else bool(v) if isinstance(v, (np.bool_,))
                else str(v) if not isinstance(v, (int, float, bool, str, type(None)))
                else v
                for v in row
            ])

    # ── Aba Por DC ───────────────────────────────────────────
    ws_dc = wb.create_sheet("Por DC")
    df_to_ws(ws_dc, r_dc, start_row=1)
    for ci, h in enumerate(r_dc.columns, 1):
        ws_dc.cell(1, ci).fill = hdr_fill
        ws_dc.cell(1, ci).font = hdr_font
        ws_dc.cell(1, ci).alignment = ctr
    for row in ws_dc.iter_rows(min_row=2):
        for cell in row:
            cell.border = brd; cell.alignment = ctr
        taxa_cell = row[4] if len(row) > 4 else None
        if taxa_cell:
            try:
                v = float(str(taxa_cell.value).replace("%","").replace(",","."))
                taxa_cell.fill = ok_fill if v >= 95 else nok_fill
                taxa_cell.font = Font(bold=True, name="Calibri", size=9,
                                       color=VERD if v >= 95 else VERM)
            except Exception:
                pass
    auto_w(ws_dc); ws_dc.freeze_panes = "A2"

    # ── Aba Base Consolidada (apenas ERROS, limitado a 100k linhas) ─
    ws_base = wb.create_sheet("Base Consolidada")
    df_exp  = df[df["OUT BOUND"] != "OK"].drop(columns=["__BASE_PAI_DST__","_ok","_er","_fa"], errors="ignore").copy()
    MAX_ROWS = 100_000
    if len(df_exp) > MAX_ROWS:
        log_cb(f"   ⚠️ {len(df_exp):,} erros — limitando a {MAX_ROWS:,} linhas no Excel para evitar travamento")
        df_exp = df_exp.head(MAX_ROWS)
    else:
        log_cb(f"   📋 {len(df_exp):,} registros com erro/fora abrangência na aba Base Consolidada")
    df_to_ws(ws_base, df_exp, start_row=1)
    fmt_hdr(ws_base)
    try:
        hdrs   = [ws_base.cell(1, ci).value for ci in range(1, ws_base.max_column+1)]
        ob_idx = hdrs.index("OUT BOUND") + 1
        er_b   = PatternFill("solid", fgColor="FFE0E0")
        fa_b   = PatternFill("solid", fgColor="FFF3CD")
        # Só colore se < 50k linhas para evitar crash
        if ws_base.max_row < 50000:
            for row in ws_base.iter_rows(min_row=2):
                val  = str(row[ob_idx-1].value or "").strip().upper()
                fill = (er_b if val == "ERRO EXPEDIÇÃO" else
                        fa_b if val == "FORA ABRANGÊNCIA" else None)
                if fill:
                    for cell in row: cell.fill = fill
    except Exception:
        pass
    auto_w(ws_base)

    # ── Outras abas ──────────────────────────────────────────
    for name, df_s in [("Por LoadingScan", r_arq),
                        ("Por Supervisor",  r_sup),
                        ("Por Tipo",        r_tipo)]:
        if df_s is not None and len(df_s):
            ws = wb.create_sheet(name)
            df_to_ws(ws, df_s, start_row=1)
            fmt_hdr(ws); auto_w(ws)

    # ── Dashboard ─────────────────────────────────────────────
    ws_dash.sheet_view.showGridLines = False
    ws_dash.merge_cells("A1:J1")
    t = ws_dash["A1"]
    t.value     = f"ERRO DE TRIAGEM (DC > DS)  —  {datetime.now().strftime('%d/%m/%Y')}"
    t.font      = Font(bold=True, color=BRANCO, name="Calibri", size=16)
    t.fill      = PatternFill("solid", fgColor=AZUL_ESC)
    t.alignment = ctr
    ws_dash.row_dimensions[1].height = 36

    headers = [COL_LC, "Triagem OK", "Triagem NOK", "Total Expedido", "Taxa (%)"]
    for ci, h in enumerate(headers, 1):
        cell = ws_dash.cell(3, ci, h)
        cell.fill = PatternFill("solid", fgColor=AZUL_ESC)
        cell.font = Font(bold=True, color=BRANCO, name="Calibri", size=9)
        cell.alignment = ctr
    ws_dash.row_dimensions[3].height = 28

    totais = {"OK": 0, "NOK": 0, "Total": 0}
    for ri, (_, row) in enumerate(r_dc.iterrows(), start=4):
        ok_v  = int(row.get("Triagem OK", 0))
        nok_v = int(row.get("Triagem NOK", 0))
        tot_v = int(row.get("Total Expedido", 0))
        taxa_v = row.get("Taxa (%)", 0)
        totais["OK"] += ok_v; totais["NOK"] += nok_v; totais["Total"] += tot_v
        vals = [str(row.get(COL_LC,"—")), ok_v, nok_v, tot_v, f"{taxa_v:.2f}%"]
        for ci, val in enumerate(vals, 1):
            cell = ws_dash.cell(ri, ci, val)
            cell.alignment = ctr; cell.border = brd
            cell.font = Font(name="Calibri", size=9)
            cell.fill = PatternFill("solid", fgColor=CINZA if ri%2==0 else BRANCO)
            if ci == 5:
                try:
                    v = float(str(val).replace("%","").replace(",","."))
                    cell.fill = PatternFill("solid", fgColor="D6F0DC" if v>=95 else "FFE0E0")
                    cell.font = Font(bold=True, name="Calibri", size=9,
                                     color=VERD if v>=95 else VERM)
                except Exception: pass

    last_row = 4 + len(r_dc)
    taxa_g   = round(totais["OK"]/totais["Total"]*100,2) if totais["Total"] else 0
    for ci, val in enumerate(["Total Geral", totais["OK"], totais["NOK"],
                               totais["Total"], f"{taxa_g:.2f}%"], 1):
        cell = ws_dash.cell(last_row, ci, val)
        cell.fill = PatternFill("solid", fgColor=AZUL_ESC)
        cell.font = Font(bold=True, color=BRANCO, name="Calibri", size=9)
        cell.alignment = ctr

    # Card total erro
    cc = len(headers) + 2
    ws_dash.merge_cells(start_row=3, end_row=3, start_column=cc, end_column=cc+1)
    c = ws_dash.cell(3, cc, "TOTAL ERRO")
    c.fill = PatternFill("solid", fgColor=AZUL_ESC)
    c.font = Font(bold=True, color=BRANCO, name="Calibri", size=10)
    c.alignment = ctr
    ws_dash.merge_cells(start_row=4, end_row=6, start_column=cc, end_column=cc+1)
    c2 = ws_dash.cell(4, cc, int(qtd_erro))
    c2.fill = PatternFill("solid", fgColor=VERM)
    c2.font = Font(bold=True, color=BRANCO, name="Calibri", size=28)
    c2.alignment = ctr

    # Top 5 DS
    if len(top5_ds):
        t5r = last_row + 3
        ws_dash.merge_cells(start_row=t5r, end_row=t5r, start_column=1, end_column=6)
        th = ws_dash.cell(t5r, 1, "TOP 5 DS COM MAIS ERROS DE TRIAGEM")
        th.fill = PatternFill("solid", fgColor=VERM)
        th.font = Font(bold=True, color=BRANCO, name="Calibri", size=12)
        th.alignment = ctr
        ws_dash.row_dimensions[t5r].height = 24
        for ci, h in enumerate(["DS","Total Erros"], 1):
            c = ws_dash.cell(t5r+1, ci, h)
            c.fill = PatternFill("solid", fgColor=AZUL_ESC)
            c.font = Font(bold=True, color=BRANCO, name="Calibri", size=9)
            c.alignment = ctr
        for ri2, (_, row2) in enumerate(top5_ds.iterrows(), start=t5r+2):
            ws_dash.cell(ri2, 1, str(row2["DS"])).alignment = Alignment(horizontal="left",vertical="center")
            ws_dash.cell(ri2, 2, int(row2["Total Erros"])).alignment = ctr
            for ci in [1,2]:
                ws_dash.cell(ri2,ci).border = brd
                ws_dash.cell(ri2,ci).font = Font(name="Calibri", size=9)

        # Gráfico
        chart = BarChart()
        chart.type = "col"; chart.style = 10; chart.title = None
        data_ref = Reference(ws_dash, min_col=2, max_col=2,
                              min_row=t5r+1, max_row=t5r+1+len(top5_ds))
        cats_ref = Reference(ws_dash, min_col=1,
                              min_row=t5r+2, max_row=t5r+1+len(top5_ds))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = AZUL_MED
        chart.width = 22; chart.height = 14
        ws_dash.add_chart(chart, f"C{t5r}")

    ws_dash.column_dimensions["A"].width = 20
    for ci in range(2, len(headers)+2):
        ws_dash.column_dimensions[get_column_letter(ci)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════
#  PÁGINA STREAMLIT
# ══════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════
#  VIEW: USUÁRIOS COMUNS — lê do banco
# ══════════════════════════════════════════════════════════════
def _render_triagem_viewer():
    import streamlit as st
    import pandas as pd
    from database import get_supabase

    sb = get_supabase()

    # Carrega lista de uploads disponíveis
    uploads = sb.table("triagem_uploads").select("id,data_ref,criado_por,total,qtd_ok,qtd_erro,taxa") \
                .order("criado_em", desc=True).limit(30).execute().data

    if not uploads:
        st.info("📭 Nenhum resultado disponível ainda. Aguarde o administrador processar os arquivos.")
        return

    # Seletor de data
    opcoes = {f"{u['data_ref']} — {u['qtd_ok']}/{u['total']} OK ({u['taxa']}%)": u["id"] for u in uploads}
    escolha = st.selectbox("📅 Selecionar processamento", list(opcoes.keys()))
    upload_id = opcoes[escolha]
    upload = next(u for u in uploads if u["id"] == upload_id)

    # Métricas
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Expedido", f"{upload['total']:,}")
    col2.metric("Triagem OK",     f"{upload['qtd_ok']:,}")
    col3.metric("Erros",          f"{upload['qtd_erro']:,}")
    col4.metric("Taxa OK",        f"{upload['taxa']}%")

    # Por DS
    r_ds = sb.table("triagem_por_ds").select("*").eq("upload_id", upload_id).execute().data
    if r_ds:
        st.markdown("### 📊 Resultado por DS")
        df_ds = pd.DataFrame(r_ds).drop(columns=["id","upload_id"], errors="ignore")
        df_ds.columns = [c.upper() for c in df_ds.columns]
        st.dataframe(df_ds, width="stretch")

    # Top 5
    top5 = sb.table("triagem_top5").select("*").eq("upload_id", upload_id).order("total_erros", desc=True).execute().data
    if top5:
        st.markdown("### ⚠️ Top 5 DS com mais erros")
        df_top = pd.DataFrame(top5).drop(columns=["id","upload_id"], errors="ignore")
        st.dataframe(df_top, width="stretch")

    # Por supervisor
    r_sup = sb.table("triagem_por_supervisor").select("*").eq("upload_id", upload_id).execute().data
    if r_sup:
        st.markdown("### 👥 Resultado por Supervisor")
        df_sup = pd.DataFrame(r_sup).drop(columns=["id","upload_id"], errors="ignore")
        st.dataframe(df_sup, width="stretch")

    # Download Excel
    if r_ds:
        import io
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Por DS"
        df_ds_exp = pd.DataFrame(r_ds).drop(columns=["id","upload_id"], errors="ignore")
        ws.append(list(df_ds_exp.columns))
        for row in df_ds_exp.itertuples(index=False):
            ws.append(list(row))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(
            "⬇️ Baixar Excel",
            data=buf,
            file_name=f"triagem_{upload['data_ref']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch"
        )

def render(usuario: str, is_admin: bool = False):
    import streamlit as st

    st.markdown("""
    <div class="app-header">
      <div class="app-header-icon">🔀</div>
      <div><h1>Triagem Analyzer</h1>
      <p>Análise de Erros de Triagem DC × DS — OUT BOUND</p></div>
    </div>""", unsafe_allow_html=True)

    # Usuários não-admin veem apenas o último resultado do banco
    if not is_admin:
        _render_triagem_viewer()
        return

    # ── Uploads ───────────────────────────────────────────────
    st.markdown('<div class="section-label">Arquivos de entrada</div>',
                unsafe_allow_html=True)
    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.markdown('<div class="upload-label">Loading Scan(s) <span class="badge-r">OBRIGATÓRIO</span></div>'
                    '<div class="upload-hint">1 ou mais arquivos — Waybill No. | Loading station | Destination Statio | Delivery Station</div>',
                    unsafe_allow_html=True)
        f_scans = st.file_uploader("scans", type=["xlsx","xls"], key="tr_scans",
                                    accept_multiple_files=True, label_visibility="collapsed")

    with col2:
        st.markdown('<div class="upload-label">Arquivo Bases <span class="badge-r">OBRIGATÓRIO</span></div>'
                    '<div class="upload-hint">aba "Base Erro Exp" — colunas: SUPERVISOR | BASE | BASE_PAI</div>',
                    unsafe_allow_html=True)
        f_bases = st.file_uploader("bases", type=["xlsx","xls"], key="tr_bases",
                                    label_visibility="collapsed")

    # ── Botão ─────────────────────────────────────────────────
    st.markdown('<div class="section-label">Executar</div>', unsafe_allow_html=True)
    pronto = bool(f_scans and f_bases)
    if not pronto:
        st.info("Faça o upload dos 2 arquivos obrigatórios.")

    # Aviso de arquivos grandes
    if f_scans:
        tamanho_total = sum(getattr(f, "size", 0) for f in f_scans) / (1024*1024)
        if tamanho_total > 30:
            st.warning(f"⚠️ Arquivos grandes ({tamanho_total:.0f} MB total). O processamento pode demorar alguns minutos.")
        elif tamanho_total > 10:
            st.info(f"📦 {tamanho_total:.0f} MB detectados. Processamento pode levar até 1 minuto.")

    # ── Estado do job em background ───────────────────────────
    if "triagem_job" not in st.session_state:
        st.session_state["triagem_job"] = {}
    job = st.session_state["triagem_job"]
    rodando = job.get("rodando", False)

    if st.button("▶  EXECUTAR ANÁLISE", disabled=(not pronto or rodando), width='stretch'):
        # Lê os bytes dos arquivos ANTES de passar para a thread
        # (Streamlit invalida os file objects fora da thread principal)
        scans_bytes = [(f.name, f.read()) for f in f_scans]
        bases_bytes = (f_bases.name, f_bases.read())

        st.session_state["triagem_job"] = {
            "rodando": True,
            "progresso": 0,
            "log": [],
            "ok": None,
            "res": None,
            "err": None,
        }
        st.session_state.pop("triagem_res", None)

        import threading, io
        job_ref = st.session_state["triagem_job"]

        def _worker(job, scans_bytes, bases_bytes):
            def log_cb(msg):
                job["log"].append(msg)

            def prog_cb(v):
                job["progresso"] = v

            scan_files = [io.BytesIO(b) for _, b in scans_bytes]
            for i, (name, _) in enumerate(scans_bytes):
                scan_files[i].name = name
            bases_io = io.BytesIO(bases_bytes[1])
            bases_io.name = bases_bytes[0]

            try:
                ok, res, err = run_analysis(scan_files, bases_io, log_cb, prog_cb)
                job["ok"]  = ok
                job["res"] = res
                job["err"] = err
            except Exception as e:
                job["ok"]  = False
                job["err"] = str(e)
            finally:
                job["rodando"] = False

        threading.Thread(target=_worker, args=(job_ref, scans_bytes, bases_bytes), daemon=True).start()
        st.rerun()

    # ── Polling enquanto processa ──────────────────────────────
    if rodando:
        prog = job.get("progresso", 0)
        st.progress(prog / 100, text=f"Processando… {prog}%")
        with st.expander("📋 Log em tempo real", expanded=True):
            st.code("\n".join(job.get("log", [])) or "Iniciando…")
        st.rerun()

    # ── Resultado quando terminar ──────────────────────────────
    if job.get("ok") is not None and not rodando:
        if not job["ok"]:
            st.error("Erro durante a análise:")
            st.code(job["err"])
        else:
            st.session_state["triagem_res"] = job["res"]
            # Salva no Supabase
            try:
                from database import get_supabase_admin
                import datetime
                sb = get_supabase_admin()
                res = job["res"]
                data_ref = datetime.date.today().isoformat()

                # Upload principal
                up = sb.table("triagem_uploads").insert({
                    "data_ref": data_ref,
                    "criado_por": usuario,
                    "total": int(res["total"]),
                    "qtd_ok": int(res["ok"]),
                    "qtd_erro": int(res["erro"]),
                    "qtd_fora": int(res["fora"]),
                    "taxa": float(res["taxa"]),
                }).execute()
                upload_id = up.data[0]["id"]

                # Por DS
                if not res["r_dc"].empty:
                    rows = []
                    for _, r in res["r_dc"].iterrows():
                        rows.append({
                            "upload_id": upload_id,
                            "ds": str(r.iloc[0]),
                            "total": int(r["Total Expedido"]),
                            "ok": int(r["Triagem OK"]),
                            "nok": int(r["Triagem NOK"]),
                            "fora": int(r["Fora Abrangência"]),
                            "taxa": float(r["Taxa (%)"]),
                        })
                    sb.table("triagem_por_ds").insert(rows).execute()

                # Top 5
                if not res["top5"].empty:
                    rows = [{"upload_id": upload_id, "ds": str(r["DS"]), "total_erros": int(r["Total Erros"])}
                            for _, r in res["top5"].iterrows()]
                    sb.table("triagem_top5").insert(rows).execute()

                # Por supervisor
                if not res["r_sup"].empty:
                    rows = []
                    for _, r in res["r_sup"].iterrows():
                        rows.append({
                            "upload_id": upload_id,
                            "supervisor": str(r.iloc[0]),
                            "total": int(r["Total Expedido"]),
                            "ok": int(r["Triagem OK"]),
                            "nok": int(r["Triagem NOK"]),
                            "fora": int(r["Fora Abrangência"]),
                            "taxa": float(r["Taxa (%)"]),
                        })
                    sb.table("triagem_por_supervisor").insert(rows).execute()

                st.success("✅ Resultado salvo no banco!")
            except Exception as e:
                st.warning(f"⚠️ Resultado gerado mas não salvo no banco: {e}")

            st.session_state["triagem_job"] = {}  # limpa job

        with st.expander("📋 Log de execução", expanded=not job.get("ok", True)):
            st.code("\n".join(job.get("log", [])))

    # ── Resultados ────────────────────────────────────────────
    res = st.session_state.get("triagem_res")
    if not res:
        return

    total    = res["total"]
    qtd_ok   = res["ok"]
    qtd_erro = res["erro"]
    qtd_fora = res["fora"]
    taxa     = res["taxa"]

    # KPIs
    st.markdown(f"""
    <div class="kpi-grid" style="grid-template-columns:repeat(4,1fr)">
      <div class="kpi-card c1">
        <div class="kpi-lbl">Total Expedido</div>
        <div class="kpi-val">{total:,}</div>
      </div>
      <div class="kpi-card c4">
        <div class="kpi-lbl">Triagem OK</div>
        <div class="kpi-val">{qtd_ok:,}</div>
        <div class="kpi-sub">taxa {taxa:.1f}%</div>
      </div>
      <div class="kpi-card c5">
        <div class="kpi-lbl">Erro Expedição</div>
        <div class="kpi-val">{qtd_erro:,}</div>
      </div>
      <div class="kpi-card c2">
        <div class="kpi-lbl">Fora Abrangência</div>
        <div class="kpi-val">{qtd_fora:,}</div>
      </div>
    </div>""", unsafe_allow_html=True)

    # Gráficos
    import plotly.express as px
    import plotly.graph_objects as go

    col_g1, col_g2 = st.columns(2)

    with col_g1:
        st.markdown('<div class="section-label">Por DC</div>', unsafe_allow_html=True)
        r_dc = res["r_dc"].copy()
        fig  = go.Figure()
        fig.add_trace(go.Bar(name="OK",  x=r_dc[COL_LC], y=r_dc["Triagem OK"],
                             marker_color="#10b981"))
        fig.add_trace(go.Bar(name="NOK", x=r_dc[COL_LC], y=r_dc["Triagem NOK"],
                             marker_color="#ef4444"))
        fig.add_trace(go.Scatter(name="Taxa %", x=r_dc[COL_LC],
                                  y=r_dc["Taxa (%)"], mode="lines+markers+text",
                                  yaxis="y2", line=dict(color="#2563eb", width=3),
                                  marker=dict(size=10),
                                  text=[f"{t:.1f}%" for t in r_dc["Taxa (%)"]],
                                  textposition="top center"))
        fig.update_layout(
            barmode="group", height=380,
            paper_bgcolor="#ffffff", plot_bgcolor="#f8fafc",
            font=dict(color="#1a1f2e", family="Plus Jakarta Sans"),
            yaxis2=dict(overlaying="y", side="right", ticksuffix="%", range=[0,120]),
            legend=dict(bgcolor="#f8fafc"),
            xaxis=dict(tickangle=-30, gridcolor="#e2e8f0"),
            yaxis=dict(gridcolor="#e2e8f0"),
            margin=dict(t=30, b=40, l=50, r=50))
        st.plotly_chart(fig, width='stretch')

    with col_g2:
        st.markdown('<div class="section-label">Top 5 DS com mais erros</div>',
                    unsafe_allow_html=True)
        top5 = res["top5"]
        if len(top5):
            fig2 = px.bar(top5.sort_values("Total Erros"),
                           x="Total Erros", y="DS", orientation="h",
                           color="Total Erros",
                           color_continuous_scale=["#fef2f2","#ef4444","#7f1d1d"])
            fig2.update_layout(
                height=380, showlegend=False,
                paper_bgcolor="#ffffff", plot_bgcolor="#f8fafc",
                font=dict(color="#1a1f2e"),
                xaxis=dict(gridcolor="#e2e8f0"),
                yaxis=dict(gridcolor="#e2e8f0"),
                coloraxis_showscale=False,
                margin=dict(t=30, b=40, l=10, r=30))
            st.plotly_chart(fig2, width='stretch')
        else:
            st.success("Nenhum erro de expedição encontrado! 🎉")

    # Tabelas
    st.markdown('<div class="section-label">Resultado por LoadingScan</div>',
                unsafe_allow_html=True)
    r_arq = res["r_arq"].copy()
    r_arq["Taxa (%)"] = r_arq["Taxa (%)"].map("{:.1f}%".format)
    st.dataframe(r_arq, width='stretch', hide_index=True)

    if len(res.get("r_sup", pd.DataFrame())):
        st.markdown('<div class="section-label">Por Supervisor</div>',
                    unsafe_allow_html=True)
        r_sup = res["r_sup"].copy()
        r_sup["Taxa (%)"] = r_sup["Taxa (%)"].map("{:.1f}%".format)
        st.dataframe(r_sup, width='stretch', hide_index=True)

    # Download Excel
    st.markdown('<div class="section-label">Download</div>', unsafe_allow_html=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    st.download_button(
        label="⬇️  Baixar Relatório Excel",
        data=res["excel_bytes"],
        file_name=f"Relatorio_Triagem_{ts}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width='stretch',
    )

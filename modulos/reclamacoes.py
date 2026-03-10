"""
modulos/reclamacoes.py
Pipeline de Análise de Tickets de Reclamação Logística
Adaptado para rodar dentro do Streamlit (sem caminhos fixos)
"""

import io
import warnings
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Configurações de colunas ──────────────────────────────────
GESTAO_COL_SIGLA      = "SIGLA"
GESTAO_COL_SUPERVISOR = "SUPERVISOR"
BILHETE_COL_WAYBILL   = "Associated Waybill"
CARTA_COL_CHAVE_IDX   = 0
CARTA_COL_NOME_IDX    = 37
CARTA_COL_ID_IDX      = 38
DELIVERED_COL_STATION_IDX = 15


# ══════════════════════════════════════════════════════════════
#  ETAPAS DO PIPELINE
# ══════════════════════════════════════════════════════════════

def carregar_bilhete(file) -> pd.DataFrame:
    df = pd.read_excel(file)
    for col in [c for c in df.columns if c != "Create Time"]:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).replace("nan", "")

    if "Create Time" not in df.columns:
        raise KeyError("Coluna 'Create Time' não encontrada no Bilhete.")

    def parse_ct(val):
        if pd.isna(val): return pd.NaT
        if isinstance(val, (int, float)):
            try: return pd.Timestamp("1899-12-30") + pd.Timedelta(days=float(val))
            except: return pd.NaT
        return pd.to_datetime(val, errors="coerce", dayfirst=True)

    df["Create Time"] = df["Create Time"].apply(parse_ct)
    df["Data"]  = pd.to_datetime(df["Create Time"].dt.date)
    df["Hora"]  = df["Create Time"].dt.time
    df["Week"]  = df["Create Time"].dt.isocalendar().week.astype("Int64")
    return df


def adicionar_supervisor(df: pd.DataFrame, file_gestao) -> pd.DataFrame:
    gestao = pd.read_excel(file_gestao, dtype=str)
    if GESTAO_COL_SIGLA not in gestao.columns or GESTAO_COL_SUPERVISOR not in gestao.columns:
        raise KeyError(f"Colunas '{GESTAO_COL_SIGLA}' ou '{GESTAO_COL_SUPERVISOR}' não encontradas em Gestão de Bases.")
    if "Inventory Station" not in df.columns:
        raise KeyError("Coluna 'Inventory Station' não encontrada no Bilhete.")

    mapa = (gestao[[GESTAO_COL_SIGLA, GESTAO_COL_SUPERVISOR]]
            .drop_duplicates(subset=GESTAO_COL_SIGLA)
            .set_index(GESTAO_COL_SIGLA)[GESTAO_COL_SUPERVISOR]
            .to_dict())

    df["_inv_upper"] = df["Inventory Station"].str.strip().str.upper()
    mapa_upper = {k.strip().upper(): v for k, v in mapa.items()}
    df["Supervisor"] = df["_inv_upper"].map(mapa_upper)
    df.drop(columns=["_inv_upper"], inplace=True)
    return df


def criar_colunas_auxiliares(df: pd.DataFrame) -> pd.DataFrame:
    df["Motorista"]    = None
    df["ID Motorista"] = None
    if "Delivery Station" not in df.columns:
        df["Delivery Station"] = None
    df["DS Referencia"] = df.get("Inventory Station")
    df["Status"] = None
    return df


def cruzar_carta_porte(df: pd.DataFrame, file_carta) -> pd.DataFrame:
    carta = pd.read_excel(file_carta, dtype=str)
    cols  = carta.columns.tolist()
    if len(cols) <= CARTA_COL_ID_IDX:
        raise KeyError(f"Carta de Porte tem apenas {len(cols)} colunas — esperado pelo menos {CARTA_COL_ID_IDX+1}.")

    col_chave = cols[CARTA_COL_CHAVE_IDX]
    col_nome  = cols[CARTA_COL_NOME_IDX]
    col_id    = cols[CARTA_COL_ID_IDX]

    if BILHETE_COL_WAYBILL not in df.columns:
        raise KeyError(f"Coluna '{BILHETE_COL_WAYBILL}' não encontrada no Bilhete.")

    carta[col_chave] = carta[col_chave].astype(str).str.strip()
    df[BILHETE_COL_WAYBILL] = df[BILHETE_COL_WAYBILL].astype(str).str.strip()

    mapa_nome = (carta[[col_chave, col_nome]].drop_duplicates(subset=col_chave)
                 .set_index(col_chave)[col_nome].to_dict())
    mapa_id   = (carta[[col_chave, col_id]].drop_duplicates(subset=col_chave)
                 .set_index(col_chave)[col_id].to_dict())

    df["Motorista"]    = df[BILHETE_COL_WAYBILL].map(mapa_nome)
    df["ID Motorista"] = df[BILHETE_COL_WAYBILL].map(mapa_id)
    return df


def limpar_dados(df: pd.DataFrame) -> pd.DataFrame:
    # Nenhum subtipo removido — mantém todos
    return df


def separar_periodo(df: pd.DataFrame):
    hoje     = datetime.now().date()
    data_ref = hoje - timedelta(days=1)
    inicio_mes = pd.Timestamp(data_ref.replace(day=1))
    fim_dia    = pd.Timestamp(data_ref).replace(hour=23, minute=59, second=59)

    df_dia = df.copy()
    datas_validas = df["Data"].notna().sum()
    if datas_validas > 0:
        df_mes = df[(df["Data"] >= inicio_mes) & (df["Data"] <= fim_dia)].copy()
    else:
        df_mes = df.copy()

    return df_dia, df_mes, data_ref


def agregar_por_supervisor(df_dia, df_mes) -> pd.DataFrame:
    col = _col_waybill(df_dia)
    a1 = df_dia.groupby("Supervisor")[col].count().reset_index(name="Qtd Dia")
    a2 = df_mes.groupby("Supervisor")[col].count().reset_index(name="Qtd Mês")
    r  = a1.merge(a2, on="Supervisor", how="outer").fillna(0)
    r["Qtd Dia"] = r["Qtd Dia"].astype(int)
    r["Qtd Mês"] = r["Qtd Mês"].astype(int)
    return r.sort_values("Qtd Dia", ascending=False)


def agregar_por_station(df_dia, df_mes) -> pd.DataFrame:
    col = _col_waybill(df_dia)
    a1 = df_dia.groupby("Inventory Station")[col].count().reset_index(name="Qtd Dia")
    a2 = df_mes.groupby("Inventory Station")[col].count().reset_index(name="Qtd Mês")
    r  = a1.merge(a2, on="Inventory Station", how="outer").fillna(0)
    r["Qtd Dia"] = r["Qtd Dia"].astype(int)
    r["Qtd Mês"] = r["Qtd Mês"].astype(int)
    return r.sort_values("Qtd Dia", ascending=False)


def top5_motoristas(df_dia, inativos: list = None) -> pd.DataFrame:
    """
    Retorna top 5 motoristas com mais reclamações.
    inativos: lista de id_motorista a excluir (marcados como inativos no sistema).
    """
    col = _col_waybill(df_dia)
    df_w = df_dia.copy()
    df_w["Motorista"]    = df_w["Motorista"].fillna(df_w["ID Motorista"]).fillna("Não identificado")
    df_w["ID Motorista"] = df_w["ID Motorista"].fillna("N/A")

    # Filtrar inativos
    if inativos:
        df_w = df_w[~df_w["ID Motorista"].isin(inativos)]

    r = (df_w.groupby(["Motorista","ID Motorista"])[col]
         .count().reset_index(name="Qtd Reclamações"))
    return r.sort_values("Qtd Reclamações", ascending=False).head(5).reset_index(drop=True)


def carregar_delivered(file_delivered, file_gestao) -> tuple:
    delivered = pd.read_excel(file_delivered, dtype=str)
    cols = delivered.columns.tolist()

    col_station = cols[DELIVERED_COL_STATION_IDX] if len(cols) > DELIVERED_COL_STATION_IDX else None
    col_contagem = next((c for c in cols if any(k in c.lower() for k in ["waybill","order no"])), cols[0])

    if col_station:
        est = delivered.groupby(col_station)[col_contagem].count().reset_index()
        est.columns = ["Delivery Station","Total Entregas"]
        est.sort_values("Total Entregas", ascending=False, inplace=True)
    else:
        est = pd.DataFrame(columns=["Delivery Station","Total Entregas"])

    # Supervisor via gestao_bases
    esup = pd.DataFrame(columns=["Supervisor","Total Entregas"])
    if col_station and file_gestao is not None:
        try:
            gestao = pd.read_excel(file_gestao, dtype=str)
            if GESTAO_COL_SIGLA in gestao.columns and GESTAO_COL_SUPERVISOR in gestao.columns:
                mapa = (gestao[[GESTAO_COL_SIGLA, GESTAO_COL_SUPERVISOR]]
                        .drop_duplicates(GESTAO_COL_SIGLA)
                        .set_index(GESTAO_COL_SIGLA)[GESTAO_COL_SUPERVISOR].to_dict())
                delivered["Supervisor"] = delivered[col_station].map(mapa)
                esup = (delivered[delivered["Supervisor"].notna()]
                        .groupby("Supervisor")[col_contagem].count().reset_index())
                esup.columns = ["Supervisor","Total Entregas"]
        except Exception:
            pass

    return est, esup


def _col_waybill(df):
    c = [x for x in df.columns if "waybill" in x.lower()]
    return c[0] if c else df.columns[0]


# ══════════════════════════════════════════════════════════════
#  GERAÇÃO DO EXCEL
# ══════════════════════════════════════════════════════════════

def gerar_excel(df_base, agg_sup, agg_sta, top5,
                entregas_sta, entregas_sup, data_ref,
                df_delivered_raw) -> bytes:

    output = io.BytesIO()

    # Aba 1: Ticket de reclamação
    df_exp = df_base.copy()
    if "Create Time" in df_exp.columns:
        df_exp["Create Time"] = pd.to_datetime(df_exp["Create Time"], errors="coerce").dt.date
    if "Hora" in df_exp.columns:
        df_exp["Hora"] = df_exp["Hora"].astype(str)
    prioridade = ["Order Number","Associated Waybill","Create Time","Week"]
    outras = [c for c in df_exp.columns if c not in prioridade]
    df_exp = df_exp[[c for c in prioridade if c in df_exp.columns] + outras]

    # Aba 2: Acumulado Entregas + Aba 3: TOP Ofensores (placeholder)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exp.to_excel(writer, sheet_name="Ticket de reclamação", index=False)
        if df_delivered_raw is not None:
            df_delivered_raw.to_excel(writer, sheet_name="Acumulado entregas", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="TOP Ofensores", index=False)

    output.seek(0)
    wb = load_workbook(output)
    _formatar_abas_simples(wb)
    _montar_top_ofensores(wb, df_base, entregas_sta, entregas_sup, top5, data_ref)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _formatar_abas_simples(wb):
    C_HDR = "1F4E79"
    borda = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"),  bottom=Side(style="thin"))
    C_ALT = "D9E1F2"

    for sn in ["Ticket de reclamação", "Acumulado entregas"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hf    = PatternFill("solid", fgColor=C_HDR)
        hfont = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        af    = PatternFill("solid", fgColor=C_ALT)
        for cell in ws[1]:
            cell.fill = hf; cell.font = hfont; cell.border = borda
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = borda
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if ri % 2 == 0:
                    cell.fill = af
        for col in ws.columns:
            l  = get_column_letter(col[0].column)
            ml = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[l].width = min(ml + 3, 45)
        ws.freeze_panes = "A2"


def _montar_top_ofensores(wb, df_base, entregas_sta, entregas_sup, top5, data_ref):
    tbl_ds, tbl_sup, tbl_mot = _construir_tabelas(df_base, entregas_sta, entregas_sup, top5, data_ref)

    C_TITULO  = "2F5597"
    C_HDR_DS  = "1F4E79"
    C_HDR_SUP = "375623"
    C_HDR_MOT = "C55A11"
    C_ALT     = "D9E1F2"
    C_TOTAL   = "BDD7EE"
    borda = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"),  bottom=Side(style="thin"))

    ws = wb["TOP Ofensores"]

    ncols_ds  = len(tbl_ds.columns)
    ncols_sup = len(tbl_sup.columns)
    col_sup   = ncols_ds + 2

    # Título
    fim_titulo = ncols_ds + ncols_sup + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=fim_titulo)
    t = ws.cell(1, 1)
    t.value = (f"Reclamações de Fake Delivery  |  "
               f"Referência: {data_ref.strftime('%d/%m/%Y')}  |  "
               f"Week {int(data_ref.strftime('%U')):02d}")
    t.font = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    t.fill = PatternFill("solid", fgColor=C_TITULO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    def hdr(row, col_start, colunas, cor):
        fill = PatternFill("solid", fgColor=cor)
        font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        for i, nome in enumerate(colunas):
            c = ws.cell(row, col_start+i, nome)
            c.fill = fill; c.font = font; c.border = borda
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30

    def dados(df, row_start, col_start, cor_total, cor_alt, rate_col=None):
        for ri, row_data in enumerate(df.itertuples(index=False)):
            r = row_start + ri
            is_total = any(str(v) == "TOTAL" for v in row_data)
            for ci, val in enumerate(row_data):
                c = ws.cell(r, col_start+ci, val)
                c.border = borda
                c.font   = Font(name="Calibri", size=10, bold=is_total)
                c.alignment = Alignment(horizontal="center", vertical="center")
                if is_total:
                    c.fill = PatternFill("solid", fgColor=cor_total)
                elif ri % 2 == 1:
                    c.fill = PatternFill("solid", fgColor=cor_alt)
                if rate_col is not None and ci == rate_col and not is_total:
                    try:
                        v = float(str(val).replace("%","").replace(",","."))
                        if v > 0.5:
                            c.font = Font(name="Calibri", size=10, color="FF0000", bold=True)
                    except Exception:
                        pass

    rate_idx_ds  = list(tbl_ds.columns).index("% Rate")  if "% Rate" in tbl_ds.columns  else None
    rate_idx_sup = list(tbl_sup.columns).index("% Rate") if "% Rate" in tbl_sup.columns else None

    hdr(2, 1,       tbl_ds.columns.tolist(),  C_HDR_DS)
    dados(tbl_ds,  3, 1,       C_TOTAL, C_ALT, rate_idx_ds)

    hdr(2, col_sup, tbl_sup.columns.tolist(), C_HDR_SUP)
    dados(tbl_sup, 3, col_sup, C_TOTAL, C_ALT, rate_idx_sup)

    row_mot = 3 + len(tbl_ds) + 2
    ws.merge_cells(start_row=row_mot-1, start_column=col_sup,
                   end_row=row_mot-1,   end_column=col_sup + len(tbl_mot.columns) - 1)
    tm = ws.cell(row_mot-1, col_sup)
    tm.value = "🏆 TOP Motoristas Ofensores"
    tm.font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    tm.fill  = PatternFill("solid", fgColor=C_HDR_MOT)
    tm.alignment = Alignment(horizontal="center", vertical="center")

    hdr(row_mot, col_sup, tbl_mot.columns.tolist(), C_HDR_MOT)
    dados(tbl_mot, row_mot+1, col_sup, C_TOTAL, C_ALT)

    for col in ws.columns:
        letra = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[letra].width = min(ml + 4, 35)

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def _construir_tabelas(df_base, entregas_sta, entregas_sup, top5, data_ref):
    col_wb    = _col_waybill(df_base)
    week_ref  = df_base["Week"].dropna().mode()
    week_ref  = int(week_ref.iloc[0]) if len(week_ref) > 0 else 0
    week_ant  = week_ref - 1

    inicio_mes = pd.Timestamp(data_ref.replace(day=1))
    fim_dia    = pd.Timestamp(data_ref).replace(hour=23, minute=59, second=59)

    df_dia   = df_base.copy()
    df_mes   = df_base[(df_base["Data"] >= inicio_mes) & (df_base["Data"] <= fim_dia)]
    df_wk    = df_base[df_base["Week"] == week_ref]
    df_wkant = df_base[df_base["Week"] == week_ant]

    lbl_dia  = f"Qt {data_ref.strftime('%d/%m')}"
    lbl_wka  = f"Qt Week {week_ant:02d}"
    lbl_wk   = f"Qt Week {week_ref:02d}"

    def cnt(df, grp, lbl):
        if len(df) == 0:
            return pd.DataFrame(columns=[grp, lbl])
        return df.groupby(grp)[col_wb].count().reset_index().rename(columns={col_wb: lbl})

    # ── Tabela DS ──────────────────────────────────────────────
    d1 = cnt(df_dia,   "Inventory Station", lbl_dia)
    d2 = cnt(df_wkant, "Inventory Station", lbl_wka)
    d3 = cnt(df_wk,    "Inventory Station", lbl_wk)
    d4 = cnt(df_mes,   "Inventory Station", "Qt Mês")

    tbl = (d1.merge(d2, on="Inventory Station", how="outer")
             .merge(d3, on="Inventory Station", how="outer")
             .merge(d4, on="Inventory Station", how="outer").fillna(0))

    mapa_sup = (df_base[["Inventory Station","Supervisor"]]
                .drop_duplicates("Inventory Station")
                .set_index("Inventory Station")["Supervisor"].to_dict())
    tbl.insert(0, "SUPERVISOR", tbl["Inventory Station"].map(mapa_sup).fillna("-"))
    tbl.rename(columns={"Inventory Station":"DS"}, inplace=True)

    if len(entregas_sta) > 0:
        ent = entregas_sta.rename(columns={"Delivery Station":"DS",
                                            entregas_sta.columns[-1]:"Entregas Mês"})
        tbl = tbl.merge(ent, on="DS", how="left").fillna(0)
        tbl["Entregas Mês"] = tbl["Entregas Mês"].astype(int)
        tbl["% Rate"] = (tbl["Qt Mês"] / tbl["Entregas Mês"].replace(0, float("nan"))) * 100
        tbl["% Rate"] = tbl["% Rate"].map(lambda x: round(x,2) if pd.notna(x) else 0)
    else:
        tbl["Entregas Mês"] = 0
        tbl["% Rate"] = 0

    for c in [lbl_dia, lbl_wka, lbl_wk, "Qt Mês"]:
        if c in tbl.columns:
            tbl[c] = tbl[c].astype(int)

    cols_ord = ["DS","SUPERVISOR", lbl_dia, lbl_wka, lbl_wk, "Entregas Mês","Qt Mês","% Rate"]
    tbl = tbl[[c for c in cols_ord if c in tbl.columns]]
    tbl.sort_values(lbl_dia, ascending=False, inplace=True)

    total = {c: int(tbl[c].sum()) if tbl[c].dtype in ["int64","float64"] else "TOTAL"
             for c in tbl.columns}
    if "% Rate" in total and total.get("Entregas Mês",0) > 0:
        total["% Rate"] = round(total["Qt Mês"] / total["Entregas Mês"] * 100, 2)
    tbl_ds = pd.concat([tbl, pd.DataFrame([total])], ignore_index=True)

    # ── Tabela Supervisor ──────────────────────────────────────
    s1 = cnt(df_dia,   "Supervisor", lbl_dia)
    s2 = cnt(df_wkant, "Supervisor", lbl_wka)
    s3 = cnt(df_wk,    "Supervisor", lbl_wk)
    s4 = cnt(df_mes,   "Supervisor", "Qt Mês")

    ts = (s1.merge(s2, on="Supervisor", how="outer")
            .merge(s3, on="Supervisor", how="outer")
            .merge(s4, on="Supervisor", how="outer").fillna(0))

    if len(entregas_sup) > 0:
        ent_s = entregas_sup.rename(columns={entregas_sup.columns[-1]:"Entregas Mês"})
        ts = ts.merge(ent_s, on="Supervisor", how="left").fillna(0)
        ts["Entregas Mês"] = ts["Entregas Mês"].astype(int)
        ts["% Rate"] = (ts["Qt Mês"] / ts["Entregas Mês"].replace(0, float("nan"))) * 100
        ts["% Rate"] = ts["% Rate"].map(lambda x: round(x,2) if pd.notna(x) else 0)
    else:
        ts["Entregas Mês"] = 0
        ts["% Rate"] = 0

    for c in [lbl_dia, lbl_wka, lbl_wk, "Qt Mês"]:
        if c in ts.columns:
            ts[c] = ts[c].astype(int)

    cols_s = ["Supervisor", lbl_dia, lbl_wka, lbl_wk, "Entregas Mês","Qt Mês","% Rate"]
    ts = ts[[c for c in cols_s if c in ts.columns]]
    ts.sort_values(lbl_dia, ascending=False, inplace=True)

    total_s = {c: int(ts[c].sum()) if ts[c].dtype in ["int64","float64"] else "TOTAL"
               for c in ts.columns}
    if "% Rate" in total_s and total_s.get("Entregas Mês",0) > 0:
        total_s["% Rate"] = round(total_s["Qt Mês"] / total_s["Entregas Mês"] * 100, 2)
    tbl_sup = pd.concat([ts, pd.DataFrame([total_s])], ignore_index=True)

    # ── Top Motoristas ─────────────────────────────────────────
    mot = (df_wk.groupby(["Supervisor","Motorista","ID Motorista","DS Referencia"])[col_wb]
           .count().reset_index(name=lbl_wk))
    mot.rename(columns={"DS Referencia":"DS"}, inplace=True)
    mot[f"Entregas Week {week_ref:02d}"] = "-"
    mot["% Rate"] = "-"
    mot.sort_values(lbl_wk, ascending=False, inplace=True)
    tbl_mot = mot.head(10).reset_index(drop=True)

    return tbl_ds, tbl_sup, tbl_mot


# ══════════════════════════════════════════════════════════════
#  INTERFACE STREAMLIT
# ══════════════════════════════════════════════════════════════

def render(is_admin=True):
    import streamlit as st
    import sys, os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from database import get_motoristas_status, upsert_motorista_status, listar_motoristas_inativos

    st.markdown("""
    <div class="app-header">
      <div class="app-header-icon">📋</div>
      <div><h1>Reclamações</h1>
      <p>Análise de Tickets de Fake Delivery · D-1</p></div>
    </div>""", unsafe_allow_html=True)

    aba1, aba2 = st.tabs(["⚙️ Processar Relatório", "🚗 Gestão de Motoristas"])

    # ── ABA 2: Gestão de Motoristas ────────────────────────────
    with aba2:
        _render_gestao_motoristas(st, get_motoristas_status,
                                   upsert_motorista_status, listar_motoristas_inativos)

    # ── ABA 1: Processar ───────────────────────────────────────
    with aba1:
        st.info("📂 Faça upload dos 4 arquivos abaixo para gerar o relatório.")

        col1, col2 = st.columns(2)
        with col1:
            f_bilhete  = st.file_uploader("📄 Bilhete de Reclamação (.xlsx)",  type=["xlsx"], key="rc_bilhete")
            f_carta    = st.file_uploader("📄 Consulta Carta de Porte (.xlsx)", type=["xlsx"], key="rc_carta")
        with col2:
            f_gestao   = st.file_uploader("📄 Gestão de Bases (.xlsx)",         type=["xlsx"], key="rc_gestao")
            f_delivered = st.file_uploader("📄 Delivered (.xlsx)",              type=["xlsx"], key="rc_delivered")

        # Preview dos arquivos carregados
        arquivos_ok = sum([f_bilhete is not None, f_carta is not None,
                           f_gestao is not None, f_delivered is not None])
        st.markdown(f"**{arquivos_ok}/4 arquivos carregados**")
        st.progress(arquivos_ok / 4)

        if arquivos_ok < 4:
            st.warning("⚠️ Faça upload de todos os 4 arquivos para processar.")
        else:
            st.success("✅ Todos os arquivos carregados! Clique em Processar.")

            # ── Estado do job em background ────────────────────
            rc_job = st.session_state.get("rc_job", {})
            rc_rodando = rc_job.get("rodando", False)

            if st.button("⚙️ Processar e gerar relatório", use_container_width=True,
                         key="rc_processar", disabled=rc_rodando):
                # Lê bytes antes de passar para thread
                files_bytes = {
                    "bilhete":   (f_bilhete.name,   f_bilhete.read()),
                    "carta":     (f_carta.name,     f_carta.read()),
                    "gestao":    (f_gestao.name,    f_gestao.read()),
                    "delivered": (f_delivered.name, f_delivered.read()),
                }
                st.session_state["rc_job"] = {
                    "rodando": True,
                    "etapa":   "Iniciando…",
                    "ok":      None,
                    "res":     None,
                    "err":     None,
                }

                def _worker_rc(files_bytes):
                    import io, traceback as tb
                    job = st.session_state["rc_job"]

                    def mk(key):
                        bio = io.BytesIO(files_bytes[key][1])
                        bio.name = files_bytes[key][0]
                        return bio

                    try:
                        job["etapa"] = "📥 Carregando Bilhete de Reclamação..."
                        inativos = listar_motoristas_inativos()
                        df = carregar_bilhete(mk("bilhete"))

                        job["etapa"] = "👤 Adicionando supervisores..."
                        df = adicionar_supervisor(df, mk("gestao"))

                        job["etapa"] = "🔧 Criando colunas auxiliares..."
                        df = criar_colunas_auxiliares(df)

                        job["etapa"] = "🚗 Cruzando com Carta de Porte..."
                        df = cruzar_carta_porte(df, mk("carta"))

                        job["etapa"] = "🧹 Limpando e separando período..."
                        df = limpar_dados(df)
                        df_dia, df_mes, data_ref = separar_periodo(df)

                        job["etapa"] = "📊 Agregando dados..."
                        agg_sup = agregar_por_supervisor(df_dia, df_mes)
                        agg_sta = agregar_por_station(df_dia, df_mes)
                        top5    = top5_motoristas(df_dia, inativos=inativos)

                        job["etapa"] = "📦 Carregando entregas..."
                        est, esup = carregar_delivered(mk("delivered"), mk("gestao"))
                        df_del_raw = pd.read_excel(io.BytesIO(files_bytes["delivered"][1]), dtype=str)

                        job["etapa"] = "💾 Gerando Excel..."
                        excel_bytes = gerar_excel(
                            df_base=df, agg_sup=agg_sup, agg_sta=agg_sta,
                            top5=top5, entregas_sta=est, entregas_sup=esup,
                            data_ref=data_ref, df_delivered_raw=df_del_raw
                        )

                        job["ok"]  = True
                        job["res"] = {
                            "excel_bytes": excel_bytes,
                            "nome_arquivo": f"reclamacoes_{data_ref.strftime('%Y%m%d')}.xlsx",
                            "n_registros": len(df),
                            "n_sup": agg_sup["Supervisor"].nunique(),
                            "n_sta": agg_sta["Inventory Station"].nunique(),
                            "n_mot": int(df["Motorista"].notna().sum()),
                            "top5": top5,
                            "agg_sup": agg_sup,
                            "agg_sta": agg_sta,
                            "inativos": inativos,
                        }
                    except Exception as e:
                        job["ok"]  = False
                        job["err"] = tb.format_exc()
                    finally:
                        job["rodando"] = False

                import threading
                threading.Thread(target=_worker_rc, args=(files_bytes,), daemon=True).start()
                st.rerun()

            # ── Polling enquanto processa ──────────────────────
            if rc_rodando:
                etapa = rc_job.get("etapa", "Processando…")
                st.info(f"⏳ {etapa}")
                st.progress(0.5)
                import time; time.sleep(1)
                st.rerun()

            # ── Resultado quando terminar ──────────────────────
            rc_res = rc_job.get("res")
            if rc_job.get("ok") is True and rc_res and not rc_rodando:
                if rc_res.get("inativos"):
                    st.info(f"ℹ️ {len(rc_res['inativos'])} motorista(s) inativo(s) excluído(s) do Top 5.")

                st.markdown("### 📊 Resumo")
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Total registros", rc_res["n_registros"])
                c2.metric("Supervisores",    rc_res["n_sup"])
                c3.metric("Stations",        rc_res["n_sta"])
                c4.metric("Motoristas ID'd", rc_res["n_mot"])

                st.markdown("### 🏆 Top 5 Motoristas")
                st.dataframe(rc_res["top5"], use_container_width=True)

                col_a, col_b = st.columns(2)
                with col_a:
                    st.markdown("**Por Supervisor (dia)**")
                    st.dataframe(rc_res["agg_sup"].head(10), use_container_width=True)
                with col_b:
                    st.markdown("**Por Station (dia)**")
                    st.dataframe(rc_res["agg_sta"].head(10), use_container_width=True)

                st.download_button(
                    label="⬇️ Baixar Relatório Excel",
                    data=rc_res["excel_bytes"],
                    file_name=rc_res["nome_arquivo"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            elif rc_job.get("ok") is False and not rc_rodando:
                st.error("❌ Erro durante o processamento:")
                with st.expander("Ver detalhes do erro"):
                    st.code(rc_job.get("err", ""))


# ══════════════════════════════════════════════════════════════
#  ABA: GESTÃO DE MOTORISTAS
# ══════════════════════════════════════════════════════════════

def _render_gestao_motoristas(st, get_status_fn, upsert_fn, listar_inativos_fn):
    st.markdown("#### 🚗 Status dos Motoristas")
    st.caption("Motoristas marcados como **Inativo** são excluídos automaticamente do Top 5 no próximo processamento.")

    # Carrega status atual
    status_map = get_status_fn()
    inativos   = [k for k, v in status_map.items() if not v.get("ativo", True)]

    col_info, col_add = st.columns([2, 1])
    with col_info:
        total = len(status_map)
        ativos   = total - len(inativos)
        st.markdown(f"**{total} motoristas cadastrados** · ✅ {ativos} ativos · ❌ {len(inativos)} inativos")

    # ── Tabela de motoristas cadastrados ──────────────────────
    if status_map:
        rows = []
        for id_mot, v in sorted(status_map.items(), key=lambda x: x[1].get("nome_motorista","") or ""):
            rows.append({
                "ID Motorista":  id_mot,
                "Nome":          v.get("nome_motorista") or "—",
                "Status":        "✅ Ativo" if v.get("ativo", True) else "❌ Inativo",
                "Motivo":        v.get("motivo") or "—",
                "Atualizado em": v.get("atualizado_em","")[:10] if v.get("atualizado_em") else "—",
                "Por":           v.get("atualizado_por") or "—",
            })
        df_status = __import__("pandas").DataFrame(rows)
        st.dataframe(df_status, use_container_width=True, hide_index=True)
    else:
        st.info("Nenhum motorista cadastrado ainda. Adicione abaixo.")

    st.markdown("---")
    st.markdown("#### ✏️ Alterar status de um motorista")

    import streamlit as _st
    usuario = _st.session_state.get("user_email", "sistema")

    with st.form("form_motorista_status"):
        c1, c2 = st.columns(2)
        with c1:
            id_input   = st.text_input("ID do Motorista *", placeholder="Ex: DRV001")
            nome_input = st.text_input("Nome do Motorista", placeholder="Ex: João Silva")
        with c2:
            status_sel = st.radio("Status", ["✅ Ativo", "❌ Inativo"], horizontal=True)
            motivo_input = st.text_input("Motivo (opcional)",
                                          placeholder="Ex: Afastado / Desligado / Em investigação")

        submitted = st.form_submit_button("💾 Salvar", use_container_width=True)

        if submitted:
            if not id_input.strip():
                st.error("ID do Motorista é obrigatório.")
            else:
                ativo = status_sel == "✅ Ativo"
                ok, err = upsert_fn(
                    id_motorista=id_input.strip(),
                    nome=nome_input.strip() or None,
                    ativo=ativo,
                    motivo=motivo_input.strip() or None,
                    usuario=usuario
                )
                if ok:
                    st.success(f"✅ Motorista **{id_input}** marcado como **{'Ativo' if ativo else 'Inativo'}**!")
                    st.rerun()
                else:
                    st.error(f"Erro: {err}")

    # ── Ação rápida: inverter status de motorista existente ───
    if status_map:
        st.markdown("#### ⚡ Ação rápida")
        ids_lista = sorted(status_map.keys())
        id_quick = st.selectbox("Selecionar motorista cadastrado", ids_lista, key="quick_sel")
        if id_quick:
            v = status_map[id_quick]
            status_atual = "✅ Ativo" if v.get("ativo", True) else "❌ Inativo"
            novo_status  = not v.get("ativo", True)
            label_btn    = f"{'❌ Marcar como Inativo' if v.get('ativo', True) else '✅ Reativar'}"
            st.caption(f"Status atual: **{status_atual}** · Nome: {v.get('nome_motorista') or '—'}")
            if st.button(label_btn, key="btn_quick_toggle"):
                ok, err = upsert_fn(
                    id_motorista=id_quick,
                    nome=v.get("nome_motorista"),
                    ativo=novo_status,
                    motivo=v.get("motivo"),
                    usuario=usuario
                )
                if ok:
                    st.success(f"✅ Status de **{id_quick}** atualizado!")
                    st.rerun()
                else:
                    st.error(f"Erro: {err}")

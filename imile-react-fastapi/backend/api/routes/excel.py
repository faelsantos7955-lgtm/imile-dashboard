"""
api/routes/excel.py — Endpoints de download de Excel formatado
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from api.deps import get_supabase, get_current_user
import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

router = APIRouter()

# ── Estilos ───────────────────────────────────────────────────
_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_ALT_FILL = PatternFill("solid", fgColor="D9E1F2")
_CTR      = Alignment(horizontal="center", vertical="center")
_BORDA    = Border(
    left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),  bottom=Side(style="thin", color="BFBFBF"),
)


def _write_sheet(ws, df, title=None):
    """Escreve DataFrame em worksheet com formatação corporativa."""
    start = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        t = ws.cell(1, 1, title)
        t.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
        t.fill = PatternFill("solid", fgColor="2F5597")
        t.alignment = _CTR
        ws.row_dimensions[1].height = 28
        start = 3

    # Header
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(start, ci, col)
        c.fill = _HDR_FILL; c.font = _HDR_FONT; c.alignment = _CTR; c.border = _BORDA

    # Data
    for ri, row in enumerate(df.itertuples(index=False), start + 1):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.border = _BORDA; c.alignment = _CTR
            c.font = Font(name="Calibri", size=10)
            if ri % 2 == 0:
                c.fill = _ALT_FILL
            # Format percentages
            cn = df.columns[ci - 1].lower()
            if 'taxa' in cn and isinstance(val, float):
                c.number_format = '0.0%'

    # Auto width
    for col in ws.columns:
        from openpyxl.utils import get_column_letter
        cl = get_column_letter(col[0].column)
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[cl].width = min(mx + 4, 35)


def _to_stream(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════
#  DASHBOARD EXCEL
# ══════════════════════════════════════════════════════════════
@router.get("/dashboard/{data_ref}")
def excel_dashboard(data_ref: str, user: dict = Depends(get_current_user)):
    sb = get_supabase()

    # Dados do dia
    q = sb.table("expedicao_diaria").select("*").eq("data_ref", data_ref)
    if user["bases"]:
        q = q.in_("scan_station", user["bases"])
    dia = pd.DataFrame(q.execute().data or [])

    if dia.empty:
        raise HTTPException(404, "Sem dados para essa data")

    # Cidades
    qc = sb.table("expedicao_cidades").select("*").eq("data_ref", data_ref)
    if user["bases"]:
        qc = qc.in_("scan_station", user["bases"])
    cid = pd.DataFrame(qc.execute().data or [])

    wb = Workbook()

    # Aba Consolidado
    ws = wb.active; ws.title = "Consolidado"
    df_exp = dia[["scan_station", "region", "recebido", "expedido", "entregas", "taxa_exp", "taxa_ent", "meta", "atingiu_meta"]].copy()
    df_exp.columns = ["DS", "Região", "Recebido", "Expedido", "Entregas", "Taxa Exp.", "Taxa Ent.", "Meta", "Na Meta"]
    df_exp = df_exp.sort_values("Recebido", ascending=False)
    _write_sheet(ws, df_exp, f"Dashboard Expedição — {data_ref}")

    # Aba por Região
    for regiao in ["Capital", "Metropolitan", "Countryside"]:
        df_r = df_exp[df_exp["Região"].str.lower() == regiao.lower()]
        if not df_r.empty:
            ws_r = wb.create_sheet(regiao)
            _write_sheet(ws_r, df_r.reset_index(drop=True), regiao)

    # Aba Cidades (se houver)
    if not cid.empty:
        ws_c = wb.create_sheet("Por Cidade")
        df_c = cid[["scan_station", "destination_city", "recebido", "expedido", "entregas", "taxa_exp", "taxa_ent"]].copy()
        df_c.columns = ["DS", "Cidade", "Recebido", "Expedido", "Entregas", "Taxa Exp.", "Taxa Ent."]
        df_c = df_c.sort_values(["DS", "Recebido"], ascending=[True, False])
        _write_sheet(ws_c, df_c, "Dados por Cidade")

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Dashboard_{data_ref}.xlsx"},
    )


# ══════════════════════════════════════════════════════════════
#  TRIAGEM EXCEL
# ══════════════════════════════════════════════════════════════
@router.get("/triagem/{upload_id}")
def excel_triagem(upload_id: int, user: dict = Depends(get_current_user)):
    sb = get_supabase()

    upload = sb.table("triagem_uploads").select("*").eq("id", upload_id).execute()
    if not upload.data:
        raise HTTPException(404, "Upload não encontrado")
    u = upload.data[0]

    por_ds  = pd.DataFrame(sb.table("triagem_por_ds").select("*").eq("upload_id", upload_id).execute().data or [])
    top5    = pd.DataFrame(sb.table("triagem_top5").select("*").eq("upload_id", upload_id).execute().data or [])
    por_sup = pd.DataFrame(sb.table("triagem_por_supervisor").select("*").eq("upload_id", upload_id).execute().data or [])

    wb = Workbook()

    # Dashboard
    ws = wb.active; ws.title = "Resumo"
    ws.cell(1, 1, "Data:").font = Font(bold=True)
    ws.cell(1, 2, u["data_ref"])
    ws.cell(2, 1, "Total:").font = Font(bold=True)
    ws.cell(2, 2, u["total"])
    ws.cell(3, 1, "OK:").font = Font(bold=True)
    ws.cell(3, 2, u["qtd_ok"])
    ws.cell(4, 1, "Erros:").font = Font(bold=True)
    ws.cell(4, 2, u["qtd_erro"])
    ws.cell(5, 1, "Taxa:").font = Font(bold=True)
    ws.cell(5, 2, f"{u['taxa']}%")

    # Por DS
    if not por_ds.empty:
        ws2 = wb.create_sheet("Por DS")
        df_ds = por_ds[["ds", "total", "ok", "nok", "fora", "taxa"]].copy()
        df_ds.columns = ["DS", "Total", "OK", "NOK", "Fora", "Taxa (%)"]
        _write_sheet(ws2, df_ds.sort_values("Taxa (%)"), "Resultado por DS")

    # Top 5
    if not top5.empty:
        ws3 = wb.create_sheet("Top 5 Erros")
        df_t = top5[["ds", "total_erros"]].copy()
        df_t.columns = ["DS", "Total Erros"]
        _write_sheet(ws3, df_t, "Top 5 DS com mais erros")

    # Por Supervisor
    if not por_sup.empty:
        ws4 = wb.create_sheet("Por Supervisor")
        df_s = por_sup[["supervisor", "total", "ok", "nok", "fora", "taxa"]].copy()
        df_s.columns = ["Supervisor", "Total", "OK", "NOK", "Fora", "Taxa (%)"]
        _write_sheet(ws4, df_s.sort_values("Taxa (%)"), "Resultado por Supervisor")

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Triagem_{u['data_ref']}.xlsx"},
    )


# ══════════════════════════════════════════════════════════════
#  RECLAMAÇÕES EXCEL
# ══════════════════════════════════════════════════════════════
@router.get("/reclamacoes/{upload_id}")
def excel_reclamacoes(upload_id: int, user: dict = Depends(get_current_user)):
    sb = get_supabase()

    upload = sb.table("reclamacoes_uploads").select("*").eq("id", upload_id).execute()
    if not upload.data:
        raise HTTPException(404, "Upload não encontrado")
    u = upload.data[0]

    r_sup = pd.DataFrame(sb.table("reclamacoes_por_supervisor").select("*").eq("upload_id", upload_id).execute().data or [])
    r_sta = pd.DataFrame(sb.table("reclamacoes_por_station").select("*").eq("upload_id", upload_id).execute().data or [])
    top5  = pd.DataFrame(sb.table("reclamacoes_top5").select("*").eq("upload_id", upload_id).execute().data or [])

    # Filtra inativos
    inativos_res = sb.table("motoristas_status").select("id_motorista").eq("ativo", False).execute()
    inativos = [r["id_motorista"] for r in (inativos_res.data or [])]
    if inativos and not top5.empty:
        top5 = top5[~top5["motorista"].isin(inativos)].head(5)

    wb = Workbook()

    if not r_sup.empty:
        ws1 = wb.active; ws1.title = "Por Supervisor"
        df_s = r_sup[["supervisor", "dia_total", "mes_total"]].copy()
        df_s.columns = ["Supervisor", "Qtd Dia", "Qtd Mês"]
        _write_sheet(ws1, df_s.sort_values("Qtd Dia", ascending=False), "Reclamações por Supervisor")

    if not r_sta.empty:
        ws2 = wb.create_sheet("Por Station")
        df_st = r_sta[["station", "dia_total", "mes_total"]].copy()
        df_st.columns = ["Station", "Qtd Dia", "Qtd Mês"]
        _write_sheet(ws2, df_st.sort_values("Qtd Dia", ascending=False), "Reclamações por Station")

    if not top5.empty:
        ws3 = wb.create_sheet("TOP Ofensores")
        df_t = top5[["motorista", "total"]].copy()
        df_t.columns = ["Motorista", "Total Reclamações"]
        _write_sheet(ws3, df_t, "TOP Motoristas Ofensores")

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Reclamacoes_{u['data_ref']}.xlsx"},
    )
